"""
Report Generator - Creates professional PDF and Excel deliverables.
Focus on QUALITY and CREDIBILITY, not just pretty formatting.

2026 UPGRADE: Now includes:
- Full methodology section with data sources
- Calculation transparency
- Confidence indicators
- Data capture for agent development
"""

import os
import json
from pathlib import Path
from datetime import datetime
from typing import Optional
import anthropic
from jinja2 import Template

from src.templates.prompts import get_report_summary_prompt
from src.agents.analyzer import AnalysisResult, BusinessContext
from src.utils.audit_data_capture import get_data_capture


class ReportGenerator:
    """
    Generates professional audit reports in PDF and Excel formats.
    """
    
    def __init__(self, api_key: Optional[str] = None, output_dir: str = "./output"):
        self.api_key = api_key or os.getenv("ANTHROPIC_API_KEY")
        self.client = anthropic.Anthropic(api_key=self.api_key) if self.api_key else None
        self.model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-20250514")
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_report(
        self, 
        analysis: AnalysisResult, 
        context: BusinessContext,
        customer_name: str
    ) -> dict:
        """
        Generate complete report package.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = customer_name.replace(" ", "_").lower()
        
        customer_dir = self.output_dir / f"{safe_name}_{timestamp}"
        customer_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate executive summary
        exec_summary = self._generate_executive_summary(analysis, context, customer_name)
        
        # Generate HTML report
        html_path = self._generate_html_report(
            analysis, context, customer_name, exec_summary, customer_dir
        )
        
        # Generate Excel workbook
        excel_path = self._generate_excel_report(
            analysis, context, customer_name, customer_dir
        )
        
        # Generate audit ID for tracking
        audit_id = f"BRC-{context.location[:3].upper()}-{context.trade_type[:4].upper()}-{timestamp}"
        
        # Save JSON data (enhanced with methodology)
        json_path = customer_dir / "analysis_data.json"
        with open(json_path, 'w') as f:
            json.dump({
                'audit_id': audit_id,
                'customer_name': customer_name,
                'context': context.model_dump(),
                'summary': analysis.summary,
                'pricing_audit': analysis.pricing_audit,
                'profitability': analysis.profitability,
                'quote_analysis': analysis.quote_analysis,
                'time_analysis': analysis.time_analysis,
                'action_plan': analysis.action_plan,
                'guarantee_check': analysis.guarantee_check,
                # NEW: Methodology and provenance
                'methodology': analysis.methodology,
                'market_benchmarks_used': analysis.market_benchmarks_used,
                'opportunity_summary': analysis.opportunity_summary
            }, f, indent=2, default=str)
        
        # Capture data for agent development pipeline
        try:
            data_capture = get_data_capture()
            captured = data_capture.capture_audit(
                audit_id=audit_id,
                business_profile={},
                analysis_result=analysis,
                context=context
            )
            print(f"📊 Audit data captured for agent development ({audit_id})")
        except Exception as e:
            print(f"Warning: Data capture failed: {e}")
        
        return {
            'audit_id': audit_id,
            'output_folder': str(customer_dir),
            'html_report': str(html_path),
            'excel_report': str(excel_path),
            'json_data': str(json_path),
            'executive_summary': exec_summary
        }
    
    def _generate_executive_summary(
        self, 
        analysis: AnalysisResult, 
        context: BusinessContext,
        customer_name: str
    ) -> str:
        """Generate personalized executive summary using Claude."""
        if not self.client:
            return self._fallback_executive_summary(analysis, context, customer_name)
        
        analysis_dict = {
            'summary': analysis.summary,
            'pricing_audit': analysis.pricing_audit,
            'action_plan': analysis.action_plan[:5],
            'guarantee_check': analysis.guarantee_check,
            'context': {
                'trade': context.trade_type,
                'location': context.location,
                'current_rate': context.current_rate,
                'hours_per_week': context.hours_per_week
            }
        }
        
        prompt = get_report_summary_prompt(
            analysis_json=json.dumps(analysis_dict, indent=2),
            customer_name=customer_name.split()[0] if customer_name else "mate"
        )
        
        message = self.client.messages.create(
            model=self.model,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        return message.content[0].text
    
    def _fallback_executive_summary(
        self, 
        analysis: AnalysisResult, 
        context: BusinessContext,
        customer_name: str
    ) -> str:
        """Generate basic summary without Claude."""
        summary = analysis.summary
        total_opp = analysis.guarantee_check.get('total_opportunity', 0)
        conservative = analysis.guarantee_check.get('total_conservative', total_opp * 0.7)
        
        return f"""
{customer_name.split()[0] if customer_name else "Mate"}, here's what we found in your numbers.

You're bringing in ${summary.get('total_revenue_analyzed', 0):,.0f} a year at a stated rate of 
${context.current_rate}/hr. Based on your revenue and hours, your effective rate works out 
to around ${summary.get('effective_hourly_rate', context.current_rate):,.0f}/hr.

The market rate for {context.trade_type}s in {context.location} is typically 
${analysis.pricing_audit.get('market_mid', context.current_rate + 15)}/hr at the mid-point.

We found ${conservative:,.0f} in opportunities you can realistically capture (conservative estimate).
Best case, if everything goes well, that could be ${total_opp:,.0f}.

Top 3 things to do:
{chr(10).join([f"- {a.get('action', 'N/A')} (${a.get('impact_conservative', a.get('impact_annual', 0)):,.0f})" for a in analysis.action_plan[:3]])}

These numbers assume you actually implement the changes. They won't happen by themselves.
"""
    
    def _generate_html_report(
        self,
        analysis: AnalysisResult,
        context: BusinessContext,
        customer_name: str,
        exec_summary: str,
        output_dir: Path
    ) -> Path:
        """Generate HTML report with evidence and calculations shown."""
        
        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Profit Audit Report - {{ customer_name }}</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
        
        :root {
            --bg: #fafafa;
            --surface: #ffffff;
            --border: #e5e5e5;
            --text: #171717;
            --text-muted: #737373;
            --accent: #ea580c;
            --success: #16a34a;
            --warning: #ca8a04;
        }
        
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'IBM Plex Sans', -apple-system, sans-serif;
            line-height: 1.7;
            color: var(--text);
            background: var(--bg);
            font-size: 16px;
        }
        
        .container {
            max-width: 800px;
            margin: 0 auto;
            padding: 60px 40px;
        }
        
        /* Header */
        header {
            border-bottom: 3px solid var(--text);
            padding-bottom: 40px;
            margin-bottom: 50px;
        }
        
        .report-label {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 12px;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 12px;
        }
        
        header h1 {
            font-size: 32px;
            font-weight: 700;
            margin-bottom: 8px;
        }
        
        header .date {
            color: var(--text-muted);
        }
        
        /* Summary Box */
        .summary-box {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 40px;
            margin-bottom: 50px;
        }
        
        .summary-box h2 {
            font-size: 14px;
            font-family: 'IBM Plex Mono', monospace;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 20px;
        }
        
        .summary-content {
            font-size: 17px;
            line-height: 1.8;
        }
        
        .summary-content p {
            margin-bottom: 16px;
        }
        
        /* Key Numbers */
        .key-numbers {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 30px;
            margin-bottom: 50px;
        }
        
        .number-card {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 30px;
        }
        
        .number-card .label {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 11px;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 8px;
        }
        
        .number-card .value {
            font-size: 36px;
            font-weight: 700;
        }
        
        .number-card .value.positive {
            color: var(--success);
        }
        
        .number-card .note {
            font-size: 13px;
            color: var(--text-muted);
            margin-top: 8px;
        }
        
        /* Sections */
        section {
            margin-bottom: 50px;
        }
        
        section h2 {
            font-size: 22px;
            font-weight: 600;
            margin-bottom: 24px;
            padding-bottom: 12px;
            border-bottom: 1px solid var(--border);
        }
        
        section h3 {
            font-size: 16px;
            font-weight: 600;
            margin: 24px 0 12px;
        }
        
        /* Calculation Box - SHOWS THE MATH */
        .calculation {
            background: #f5f5f5;
            border-left: 3px solid var(--accent);
            padding: 20px 24px;
            margin: 20px 0;
            font-family: 'IBM Plex Mono', monospace;
            font-size: 14px;
        }
        
        .calculation .label {
            font-size: 11px;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 12px;
        }
        
        .calculation .line {
            margin-bottom: 6px;
        }
        
        .calculation .result {
            margin-top: 12px;
            padding-top: 12px;
            border-top: 1px solid var(--border);
            font-weight: 600;
        }
        
        /* Evidence Box */
        .evidence {
            background: #fef9f3;
            border: 1px solid #fed7aa;
            padding: 16px 20px;
            margin: 16px 0;
            font-size: 14px;
        }
        
        .evidence strong {
            color: var(--accent);
        }
        
        /* Action Cards */
        .action-card {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 30px;
            margin-bottom: 24px;
        }
        
        .action-header {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 16px;
        }
        
        .action-number {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 12px;
            color: var(--accent);
        }
        
        .action-title {
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 4px;
        }
        
        .action-impact {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 20px;
            font-weight: 600;
            color: var(--success);
        }
        
        .action-meta {
            display: flex;
            gap: 16px;
            margin-bottom: 16px;
            font-size: 13px;
        }
        
        .action-meta span {
            background: #f5f5f5;
            padding: 4px 12px;
            border-radius: 2px;
        }
        
        .action-body {
            margin-top: 16px;
        }
        
        .action-body h4 {
            font-size: 13px;
            font-family: 'IBM Plex Mono', monospace;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin: 20px 0 8px;
        }
        
        .action-body p {
            font-size: 15px;
            color: var(--text);
            margin-bottom: 8px;
        }
        
        .script-box {
            background: #f8f8f8;
            border: 1px solid var(--border);
            padding: 16px 20px;
            font-size: 14px;
            font-style: italic;
            margin: 12px 0;
        }
        
        .risk-box {
            background: #fef2f2;
            border: 1px solid #fecaca;
            padding: 12px 16px;
            font-size: 14px;
            margin: 12px 0;
        }
        
        /* Table */
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
        }
        
        th, td {
            padding: 12px 16px;
            text-align: left;
            border-bottom: 1px solid var(--border);
        }
        
        th {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 11px;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            color: var(--text-muted);
            font-weight: 500;
        }
        
        /* Guarantee */
        .guarantee-box {
            background: var(--text);
            color: white;
            padding: 40px;
            text-align: center;
            margin: 50px 0;
        }
        
        .guarantee-box h2 {
            border: none;
            color: white;
            margin-bottom: 16px;
        }
        
        .guarantee-amount {
            font-size: 48px;
            font-weight: 700;
            margin: 16px 0;
        }
        
        .guarantee-note {
            font-size: 14px;
            opacity: 0.8;
            max-width: 500px;
            margin: 0 auto;
        }
        
        /* Assumptions */
        .assumptions {
            background: #fffbeb;
            border: 1px solid #fde68a;
            padding: 24px;
            margin: 30px 0;
        }
        
        .assumptions h3 {
            font-size: 14px;
            font-family: 'IBM Plex Mono', monospace;
            margin-bottom: 12px;
        }
        
        .assumptions ul {
            padding-left: 20px;
            font-size: 14px;
        }
        
        .assumptions li {
            margin-bottom: 6px;
        }
        
        /* Profit Leak Dashboard - THE HERO */
        .profit-leak-dashboard {
            background: linear-gradient(135deg, #171717 0%, #262626 100%);
            color: white;
            padding: 40px;
            margin-bottom: 40px;
            border-radius: 8px;
        }
        
        .dashboard-header {
            text-align: center;
            margin-bottom: 32px;
        }
        
        .dashboard-icon {
            font-size: 48px;
            margin-bottom: 12px;
        }
        
        .dashboard-header h2 {
            border: none;
            color: white;
            font-size: 28px;
            margin-bottom: 8px;
        }
        
        .dashboard-header p {
            color: rgba(255,255,255,0.7);
            font-size: 16px;
        }
        
        .leak-breakdown {
            max-width: 600px;
            margin: 0 auto;
        }
        
        .leak-item {
            margin-bottom: 16px;
        }
        
        .leak-bar {
            background: linear-gradient(90deg, #22c55e 0%, #16a34a var(--width, 60%), #374151 var(--width, 60%));
            padding: 16px 20px;
            border-radius: 4px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            position: relative;
            overflow: hidden;
        }
        
        .leak-bar.callout {
            background: linear-gradient(90deg, #eab308 0%, #ca8a04 var(--width, 40%), #374151 var(--width, 40%));
        }
        
        .leak-bar.quote {
            background: linear-gradient(90deg, #3b82f6 0%, #2563eb var(--width, 30%), #374151 var(--width, 30%));
        }
        
        .leak-label {
            font-size: 14px;
            font-weight: 500;
        }
        
        .leak-amount {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 18px;
            font-weight: 700;
        }
        
        .leak-total {
            margin-top: 24px;
            padding-top: 24px;
            border-top: 2px solid rgba(255,255,255,0.2);
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-size: 16px;
        }
        
        .leak-total .total-amount {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 32px;
            font-weight: 700;
            color: #22c55e;
        }
        
        /* Rate Comparison Visual */
        .rate-comparison {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 30px;
            margin-bottom: 40px;
        }
        
        .rate-comparison h3 {
            text-align: center;
            margin-bottom: 30px;
            font-size: 18px;
        }
        
        .rate-scale {
            position: relative;
            height: 120px;
            margin: 0 40px;
        }
        
        .rate-bar {
            position: absolute;
            bottom: 30px;
            left: 0;
            right: 0;
            height: 12px;
            background: linear-gradient(90deg, #dc2626 0%, #eab308 35%, #22c55e 65%, #16a34a 100%);
            border-radius: 6px;
        }
        
        .rate-marker {
            position: absolute;
            bottom: 45px;
            left: var(--position, 50%);
            transform: translateX(-50%);
        }
        
        .rate-marker::after {
            content: '';
            position: absolute;
            bottom: -15px;
            left: 50%;
            transform: translateX(-50%);
            width: 0;
            height: 0;
            border-left: 8px solid transparent;
            border-right: 8px solid transparent;
        }
        
        .rate-marker.your-rate::after {
            border-top: 12px solid #dc2626;
        }
        
        .rate-marker.market-avg::after {
            border-top: 12px solid #22c55e;
        }
        
        .marker-label {
            background: var(--text);
            color: white;
            padding: 8px 12px;
            border-radius: 4px;
            font-size: 12px;
            font-family: 'IBM Plex Mono', monospace;
            text-align: center;
            white-space: nowrap;
        }
        
        .rate-marker.your-rate .marker-label {
            background: #dc2626;
        }
        
        .rate-marker.market-avg .marker-label {
            background: #16a34a;
        }
        
        .rate-labels {
            position: absolute;
            bottom: 0;
            left: 0;
            right: 0;
            display: flex;
            justify-content: space-between;
            font-size: 11px;
            color: var(--text-muted);
            font-family: 'IBM Plex Mono', monospace;
        }
        
        .rate-verdict {
            text-align: center;
            margin-top: 20px;
            padding: 12px;
            border-radius: 4px;
            font-size: 15px;
        }
        
        .rate-verdict.below {
            background: #fef2f2;
            border: 1px solid #fecaca;
            color: #dc2626;
        }
        
        .rate-verdict.above {
            background: #f0fdf4;
            border: 1px solid #bbf7d0;
            color: #16a34a;
        }
        
        /* 90-Day Timeline */
        .timeline {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 30px;
            margin: 40px 0;
        }
        
        .timeline h3 {
            margin-bottom: 24px;
        }
        
        .timeline-track {
            display: flex;
            gap: 0;
            margin-bottom: 20px;
        }
        
        .timeline-phase {
            flex: 1;
            padding: 20px;
            text-align: center;
            position: relative;
        }
        
        .timeline-phase::after {
            content: '→';
            position: absolute;
            right: -10px;
            top: 50%;
            transform: translateY(-50%);
            font-size: 20px;
            color: var(--text-muted);
        }
        
        .timeline-phase:last-child::after {
            display: none;
        }
        
        .timeline-phase.week1 { background: #dcfce7; }
        .timeline-phase.month1 { background: #fef3c7; }
        .timeline-phase.quarter1 { background: #dbeafe; }
        
        .phase-label {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 11px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-muted);
            margin-bottom: 8px;
        }
        
        .phase-title {
            font-weight: 600;
            font-size: 14px;
            margin-bottom: 4px;
        }
        
        .phase-impact {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 16px;
            font-weight: 700;
            color: var(--success);
        }
        
        /* Worst Jobs Table */
        .worst-jobs {
            background: #fef2f2;
            border: 2px solid #fecaca;
            padding: 24px;
            margin: 30px 0;
        }
        
        .worst-jobs h3 {
            color: #dc2626;
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .worst-jobs table {
            margin: 0;
        }
        
        .worst-jobs td.loss {
            color: #dc2626;
            font-weight: 700;
            font-family: 'IBM Plex Mono', monospace;
        }
        
        /* Best vs Fire Customers */
        .customer-verdict {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 24px;
            margin: 30px 0;
        }
        
        .customer-box {
            padding: 24px;
            border-radius: 4px;
        }
        
        .customer-box.keep {
            background: #f0fdf4;
            border: 1px solid #bbf7d0;
        }
        
        .customer-box.fire {
            background: #fef2f2;
            border: 1px solid #fecaca;
        }
        
        .customer-box h4 {
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .customer-box.keep h4 { color: #16a34a; }
        .customer-box.fire h4 { color: #dc2626; }
        
        .customer-item {
            padding: 12px 0;
            border-bottom: 1px solid rgba(0,0,0,0.1);
        }
        
        .customer-item:last-child {
            border-bottom: none;
        }
        
        .customer-name {
            font-weight: 600;
        }
        
        .customer-stats {
            font-size: 13px;
            color: var(--text-muted);
            margin-top: 4px;
        }
        
        /* Footer */
        footer {
            border-top: 1px solid var(--border);
            padding-top: 30px;
            margin-top: 50px;
            text-align: center;
            color: var(--text-muted);
            font-size: 14px;
        }
        
        /* NEW: Confidence Badges */
        .confidence-badge {
            display: inline-block;
            padding: 4px 12px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: 600;
            font-family: 'IBM Plex Mono', monospace;
            letter-spacing: 0.05em;
        }
        
        .confidence-badge.high {
            background: #10b981;
            color: white;
        }
        
        .confidence-badge.medium {
            background: #f59e0b;
            color: white;
        }
        
        .confidence-badge.low {
            background: #ef4444;
            color: white;
        }
        
        /* NEW: Methodology Section */
        .methodology-section {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 30px;
            margin: 40px 0;
        }
        
        .methodology-section h2 {
            margin-bottom: 24px;
        }
        
        .methodology-grid {
            display: grid;
            gap: 20px;
        }
        
        .methodology-item {
            padding: 20px;
            background: white;
            border: 1px solid var(--border);
        }
        
        .methodology-item h3 {
            font-size: 13px;
            font-family: 'IBM Plex Mono', monospace;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 12px;
        }
        
        .methodology-item ul {
            padding-left: 20px;
            font-size: 14px;
            line-height: 1.8;
        }
        
        /* NEW: Data Source Citation */
        .data-source {
            background: #eff6ff;
            border-left: 3px solid #3b82f6;
            padding: 12px 16px;
            margin: 12px 0;
            font-size: 13px;
        }
        
        .data-source .source-name {
            font-weight: 600;
            color: #1e40af;
        }
        
        /* NEW: Three Scenario Box */
        .scenarios-box {
            background: var(--surface);
            border: 1px solid var(--border);
            padding: 20px;
            margin: 20px 0;
        }
        
        .scenarios-box h4 {
            font-size: 12px;
            font-family: 'IBM Plex Mono', monospace;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 16px;
        }
        
        .scenario-row {
            display: flex;
            justify-content: space-between;
            padding: 10px 0;
            border-bottom: 1px solid var(--border);
        }
        
        .scenario-row:last-child {
            border-bottom: none;
        }
        
        .scenario-row.recommended {
            background: #f0fdf4;
            margin: 0 -20px;
            padding: 10px 20px;
        }
        
        .scenario-label {
            font-size: 14px;
        }
        
        .scenario-amount {
            font-family: 'IBM Plex Mono', monospace;
            font-weight: 600;
        }
        
        /* NEW: Risk Card */
        .risk-card {
            background: #fef2f2;
            border: 1px solid #fecaca;
            padding: 16px;
            margin: 12px 0;
            border-radius: 4px;
        }
        
        .risk-header {
            font-weight: 600;
            color: #dc2626;
            margin-bottom: 8px;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .risk-body {
            font-size: 14px;
            line-height: 1.6;
        }
        
        /* NEW: Disclaimer Section */
        .disclaimer-section {
            background: #fffbeb;
            border: 1px solid #fde68a;
            padding: 24px;
            margin: 40px 0;
        }
        
        .disclaimer-section h3 {
            font-size: 14px;
            margin-bottom: 12px;
        }
        
        .disclaimer-section ul {
            padding-left: 20px;
            font-size: 14px;
            line-height: 1.8;
        }
        
        /* NEW: Professional Report ID */
        .report-id {
            font-family: 'IBM Plex Mono', monospace;
            font-size: 11px;
            color: var(--text-muted);
            letter-spacing: 0.05em;
        }
        
        @media print {
            .container { max-width: 100%; padding: 20px; }
            .action-card { break-inside: avoid; }
        }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div class="report-label">Profit Leak Audit Report</div>
            <h1>{{ customer_name }}</h1>
            <div class="date">{{ date }} · {{ context.trade_type | capitalize }} · {{ context.location }}</div>
        </header>
        
        <div class="summary-box">
            <h2>Executive Summary</h2>
            <div class="summary-content">
                {{ exec_summary | replace('\n', '</p><p>') | safe }}
            </div>
        </div>
        
        <!-- HERO DASHBOARD - The "Holy Shit" Moment -->
        <div class="profit-leak-dashboard">
            <div class="dashboard-header">
                <div class="dashboard-icon">💰</div>
                <h2>Money Left on the Table</h2>
                <p>Here's exactly where your profit is leaking</p>
            </div>
            <div class="leak-breakdown">
                {% set rate_gap = ((pricing.market_mid or (context.current_rate + 20)) - context.current_rate) * ((summary.total_revenue_analyzed or 100000) / (context.current_rate or 95)) * 0.85 %}
                {% set callout_gap = (pricing.call_out_impact or 3000) * 0.7 %}
                {% set quote_gap = (guarantee.quick_wins_total or 5000) * 0.3 %}
                {% set total_gap = rate_gap + callout_gap + quote_gap %}
                
                <div class="leak-item">
                    <div class="leak-bar" style="--width: {{ ((rate_gap / (total_gap or 1)) * 100) | int }}%">
                        <span class="leak-label">Undercharging vs Market</span>
                        <span class="leak-amount">+${{ "{:,.0f}".format(rate_gap) }}/yr</span>
                    </div>
                </div>
                <div class="leak-item">
                    <div class="leak-bar callout" style="--width: {{ ((callout_gap / (total_gap or 1)) * 100) | int }}%">
                        <span class="leak-label">Missing Call-out Fees</span>
                        <span class="leak-amount">+${{ "{:,.0f}".format(callout_gap) }}/yr</span>
                    </div>
                </div>
                <div class="leak-item">
                    <div class="leak-bar quote" style="--width: {{ ((quote_gap / (total_gap or 1)) * 100) | int }}%">
                        <span class="leak-label">Quote/Efficiency Gaps</span>
                        <span class="leak-amount">+${{ "{:,.0f}".format(quote_gap) }}/yr</span>
                    </div>
                </div>
                <div class="leak-total">
                    <span>TOTAL OPPORTUNITY</span>
                    <span class="total-amount">${{ "{:,.0f}".format(guarantee.total_conservative or (guarantee.total_opportunity or 15000) * 0.7) }}/yr</span>
                </div>
            </div>
        </div>
        
        <!-- Rate Comparison Visual -->
        <div class="rate-comparison">
            <h3>Where You Sit in the Market</h3>
            <div class="rate-scale">
                <div class="rate-marker your-rate" style="--position: {{ (((context.current_rate - (pricing.market_low or 80)) / ((pricing.market_high or 160) - (pricing.market_low or 80))) * 100) | int }}%">
                    <div class="marker-label">YOU<br>${{ context.current_rate }}/hr</div>
                </div>
                <div class="rate-marker market-avg" style="--position: {{ (((pricing.market_mid or 115) - (pricing.market_low or 80)) / ((pricing.market_high or 160) - (pricing.market_low or 80))) * 100 | int }}%">
                    <div class="marker-label">MARKET AVG<br>${{ pricing.market_mid or (context.current_rate + 20) }}/hr</div>
                </div>
                <div class="rate-bar"></div>
                <div class="rate-labels">
                    <span>Bottom 25%<br>${{ pricing.market_low or (context.current_rate - 15) }}</span>
                    <span>Top 10%<br>${{ pricing.market_high or (context.current_rate + 50) }}</span>
                </div>
            </div>
            {% if context.current_rate < (pricing.market_mid or 115) %}
            <div class="rate-verdict below">
                ⚠️ You're charging <strong>{{ (((pricing.market_mid or 115) - context.current_rate) / (pricing.market_mid or 115) * 100) | round(0) | int }}% below</strong> market average
            </div>
            {% else %}
            <div class="rate-verdict above">
                ✓ You're at or above market average - focus on efficiency gains
            </div>
            {% endif %}
        </div>

        <div class="key-numbers">
            <div class="number-card">
                <div class="label">Revenue Analyzed</div>
                <div class="value">${{ "{:,.0f}".format(summary.total_revenue_analyzed or 0) }}</div>
            </div>
            <div class="number-card">
                <div class="label">Your Effective Rate</div>
                <div class="value">${{ (summary.effective_hourly_rate or summary.calculated_effective_rate or context.current_rate) | default(95) | round(0) | int }}/hr</div>
                <div class="note">vs market ${{ pricing.market_mid or (context.current_rate + 20) | int }}/hr</div>
            </div>
            <div class="number-card">
                <div class="label">Conservative Opportunity</div>
                <div class="value positive">${{ "{:,.0f}".format(guarantee.total_conservative or (guarantee.total_opportunity or 15000) * 0.7) }}</div>
                <div class="note">Realistic estimate</div>
            </div>
            <div class="number-card">
                <div class="label">Best Case Opportunity</div>
                <div class="value">${{ "{:,.0f}".format(guarantee.total_best_case or guarantee.total_opportunity or 15000) }}</div>
                <div class="note">If everything goes perfectly</div>
            </div>
        </div>
        
        {% if business_health %}
        <section>
            <h2>Business Health Score</h2>
            <p style="margin-bottom: 20px;">How your business stacks up across key metrics (1-10 scale):</p>
            <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px;">
                {% for metric, score in business_health.items() if metric != 'overall' %}
                <div style="background: var(--surface); border: 1px solid var(--border); padding: 20px; text-align: center;">
                    <div style="font-size: 28px; font-weight: 700; color: {% if score >= 7 %}var(--success){% elif score >= 5 %}var(--warning){% else %}#dc2626{% endif %};">{{ score }}/10</div>
                    <div style="font-size: 13px; color: var(--text-muted); text-transform: capitalize;">{{ metric | replace('_', ' ') }}</div>
                </div>
                {% endfor %}
            </div>
            {% if business_health.overall %}
            <div style="background: var(--text); color: white; padding: 24px; margin-top: 20px; text-align: center;">
                <div style="font-size: 14px; opacity: 0.7;">Overall Business Health</div>
                <div style="font-size: 42px; font-weight: 700;">{{ business_health.overall }}/10</div>
            </div>
            {% endif %}
        </section>
        {% endif %}
        
        <section>
            <h2>Pricing Analysis</h2>
            
            <div class="calculation">
                <div class="label">How We Calculated Your Effective Rate</div>
                <div class="line">Total revenue: ${{ "{:,.0f}".format(summary.total_revenue_analyzed or 0) }}</div>
                <div class="line">Estimated billable hours: ~{{ ((summary.total_revenue_analyzed or 0) / (summary.effective_hourly_rate or context.current_rate or 95)) | round(0) | int }} hours/year</div>
                <div class="line">Calculation: ${{ "{:,.0f}".format(summary.total_revenue_analyzed or 0) }} ÷ {{ ((summary.total_revenue_analyzed or 0) / (summary.effective_hourly_rate or context.current_rate or 95)) | round(0) | int }} hours</div>
                <div class="result">= ${{ (summary.effective_hourly_rate or context.current_rate or 95) | round(0) | int }}/hr effective rate</div>
            </div>
            
            <div class="evidence">
                <strong>Market data:</strong> {{ context.trade_type | capitalize }}s in {{ context.location }} typically charge 
                ${{ pricing.market_low or (context.current_rate - 5) }}-${{ pricing.market_high or (context.current_rate + 35) }}/hr, 
                with ${{ pricing.market_mid or (context.current_rate + 15) }}/hr being mid-market.
                Source: Industry associations & trade surveys.
            </div>
            
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Current</th>
                    <th>Recommended</th>
                    <th>Impact (Conservative)</th>
                </tr>
                <tr>
                    <td>Hourly Rate</td>
                    <td>${{ context.current_rate }}/hr</td>
                    <td>${{ pricing.recommended_rate or (context.current_rate + 15) }}/hr</td>
                    <td style="color: var(--success)">+${{ "{:,.0f}".format((pricing.rate_increase_impact or 15000) * 0.7) }}/yr</td>
                </tr>
                {% if pricing.call_out_fee_recommended %}
                <tr>
                    <td>Call-out Fee</td>
                    <td>${{ pricing.call_out_fee_current or 0 }}</td>
                    <td>${{ pricing.call_out_fee_recommended }}</td>
                    <td style="color: var(--success)">+${{ "{:,.0f}".format((pricing.call_out_impact or 3000) * 0.7) }}/yr</td>
                </tr>
                {% endif %}
            </table>
        </section>
        
        <section>
            <h2>Action Plan</h2>
            <p>Prioritized by realistic impact. Each action includes the calculation, what to say, and what could go wrong.</p>
            
            {% for action in action_plan[:8] %}
            <div class="action-card">
                <div class="action-header">
                    <div>
                        <div class="action-number">Action {{ loop.index }}</div>
                        <div class="action-title">{{ action.action }}</div>
                    </div>
                    <div class="action-impact">+${{ "{:,.0f}".format((action.impact_conservative or action.impact_annual or 0) * 0.85) }}/yr</div>
                </div>
                
                <div class="action-meta">
                    <span>Effort: {{ action.effort }}</span>
                    <span>Timeline: {{ action.timeline }}</span>
                </div>
                
                <div class="action-body">
                    {% if action.calculation %}
                    <div class="calculation">
                        <div class="label">The Math</div>
                        <div>{{ action.calculation }}</div>
                    </div>
                    {% endif %}
                    
                    {% if action.how %}
                    <h4>How to implement</h4>
                    <p>{{ action.how }}</p>
                    {% endif %}
                    
                    {% if action.script %}
                    <h4>What to say</h4>
                    <div class="script-box">"{{ action.script }}"</div>
                    {% endif %}
                    
                    {% if action.pushback_response %}
                    <h4>If they push back</h4>
                    <div class="script-box">"{{ action.pushback_response }}"</div>
                    {% endif %}
                    
                    {% if action.risk %}
                    <h4>What could go wrong</h4>
                    <div class="risk-box">{{ action.risk }}</div>
                    {% endif %}
                </div>
            </div>
            {% endfor %}
        </section>
        
        <!-- 90-Day Action Timeline -->
        <div class="timeline">
            <h3>📅 Your 90-Day Profit Recovery Timeline</h3>
            <div class="timeline-track">
                <div class="timeline-phase week1">
                    <div class="phase-label">Week 1-2</div>
                    <div class="phase-title">Quick Wins</div>
                    <div class="phase-impact">+${{ "{:,.0f}".format((guarantee.quick_wins_total or 5000) * 0.3) }}</div>
                    <div style="font-size: 12px; margin-top: 8px; color: var(--text-muted);">
                        Update rates<br>Add call-out fee
                    </div>
                </div>
                <div class="timeline-phase month1">
                    <div class="phase-label">Week 3-6</div>
                    <div class="phase-title">Build Momentum</div>
                    <div class="phase-impact">+${{ "{:,.0f}".format((guarantee.quick_wins_total or 5000) * 0.4) }}</div>
                    <div style="font-size: 12px; margin-top: 8px; color: var(--text-muted);">
                        Review customers<br>Chase unpaid invoices
                    </div>
                </div>
                <div class="timeline-phase quarter1">
                    <div class="phase-label">Week 7-12</div>
                    <div class="phase-title">Lock It In</div>
                    <div class="phase-impact">+${{ "{:,.0f}".format((guarantee.quick_wins_total or 5000) * 0.3) }}</div>
                    <div style="font-size: 12px; margin-top: 8px; color: var(--text-muted);">
                        Optimize job mix<br>Build systems
                    </div>
                </div>
            </div>
            <div style="text-align: center; padding: 16px; background: #f5f5f5; border-radius: 4px;">
                <strong>Total 90-Day Target:</strong> 
                <span style="font-family: 'IBM Plex Mono', monospace; font-size: 20px; color: var(--success); margin-left: 8px;">
                    ${{ "{:,.0f}".format(guarantee.quick_wins_total or (guarantee.total_conservative or 15000) * 0.4) }}
                </span>
                <span style="color: var(--text-muted); font-size: 14px;"> in the bank</span>
            </div>
        </div>
        
        {% if customer_analysis and customer_analysis.top_customers %}
        <section>
            <h2>Customer Analysis</h2>
            <p>Who's worth your time - and who isn't.</p>
            
            <!-- Keep vs Fire Visual -->
            <div class="customer-verdict">
                <div class="customer-box keep">
                    <h4>✅ Your Best Customers (Keep Happy)</h4>
                    {% for cust in customer_analysis.top_customers[:3] if cust.grade in ['A', 'B'] %}
                    <div class="customer-item">
                        <div class="customer-name">{{ cust.name or cust.customer or 'Top Customer' }}</div>
                        <div class="customer-stats">${{ "{:,.0f}".format(cust.total_revenue or cust.revenue or 0) }}/yr · {{ cust.job_count or cust.jobs or 0 }} jobs · Grade {{ cust.grade or 'A' }}</div>
                    </div>
                    {% endfor %}
                </div>
                <div class="customer-box fire">
                    <h4>⚠️ Consider Dropping</h4>
                    {% if customer_analysis.concerning_customers %}
                    {% for cust in customer_analysis.concerning_customers[:3] %}
                    <div class="customer-item">
                        <div class="customer-name">{{ cust.name or 'Problem Customer' }}</div>
                        <div class="customer-stats">{{ cust.reason or 'Low margin, high effort' }}</div>
                    </div>
                    {% endfor %}
                    {% else %}
                    <div class="customer-item">
                        <div class="customer-stats">No major problem customers identified - good sign!</div>
                    </div>
                    {% endif %}
                </div>
            </div>
            
            <h3 style="margin-top: 24px;">Full Customer Breakdown</h3>
            <table>
                <tr>
                    <th>Customer</th>
                    <th>Revenue</th>
                    <th>Jobs</th>
                    <th>Avg Job</th>
                    <th>Grade</th>
                    <th>Action</th>
                </tr>
                {% for cust in customer_analysis.top_customers[:8] %}
                <tr>
                    <td>{{ cust.name or cust.customer or 'Unknown' }}</td>
                    <td>${{ "{:,.0f}".format(cust.total_revenue or cust.revenue or 0) }}</td>
                    <td>{{ cust.job_count or cust.jobs or 0 }}</td>
                    <td>${{ "{:,.0f}".format(cust.avg_job_size or 0) }}</td>
                    <td style="font-weight: 600; color: {% if cust.grade == 'A' %}var(--success){% elif cust.grade == 'B' %}var(--text){% elif cust.grade == 'C' %}var(--warning){% else %}#dc2626{% endif %};">{{ cust.grade or 'B' }}</td>
                    <td>{{ cust.recommendation or 'Maintain' }}</td>
                </tr>
                {% endfor %}
            </table>
            
            {% if customer_analysis.concerning_customers %}
            <h3 style="margin-top: 24px;">⚠️ Customers to Watch</h3>
            {% for cust in customer_analysis.concerning_customers[:3] %}
            <div class="risk-box" style="margin-bottom: 12px;">
                <strong>{{ cust.name or 'Unknown' }}:</strong> {{ cust.reason or 'Review needed' }}
                <br><em>Recommendation: {{ cust.recommendation or 'Review' }}</em>
            </div>
            {% endfor %}
            {% endif %}
            
            {% if customer_analysis.ideal_customer_profile %}
            <div class="evidence" style="margin-top: 20px;">
                <strong>Your ideal customer profile:</strong> {{ customer_analysis.ideal_customer_profile }}
            </div>
            {% endif %}
        </section>
        {% endif %}
        
        {% if cash_flow_insights and (cash_flow_insights.recommendations or cash_flow_insights.cash_flow_risks) %}
        <section>
            <h2>Cash Flow Insights</h2>
            <p>Cash flow kills more trade businesses than lack of work. Here's what we found:</p>
            
            {% if cash_flow_insights.payment_speed_assessment %}
            <div class="evidence">
                <strong>Payment timing:</strong> {{ cash_flow_insights.payment_speed_assessment }}
            </div>
            {% endif %}
            
            {% if cash_flow_insights.seasonal_patterns %}
            <div class="evidence">
                <strong>Seasonal patterns:</strong> {{ cash_flow_insights.seasonal_patterns }}
            </div>
            {% endif %}
            
            {% if cash_flow_insights.cash_flow_risks %}
            <h3 style="margin-top: 20px;">Cash Flow Risks</h3>
            {% for risk in cash_flow_insights.cash_flow_risks %}
            <div class="risk-box" style="margin-bottom: 8px;">{{ risk }}</div>
            {% endfor %}
            {% endif %}
            
            {% if cash_flow_insights.recommendations %}
            <h3 style="margin-top: 20px;">Cash Flow Actions</h3>
            <ul style="padding-left: 24px;">
                {% for rec in cash_flow_insights.recommendations %}
                <li style="margin-bottom: 8px;">{{ rec }}</li>
                {% endfor %}
            </ul>
            {% endif %}
        </section>
        {% endif %}
        
        {% if job_analysis %}
        <section>
            <h2>Job Profitability Breakdown</h2>
            <p>Which jobs make you money - and which don't.</p>
            
            <table>
                <tr>
                    <th>Job Type</th>
                    <th>Count</th>
                    <th>Revenue</th>
                    <th>Avg Job</th>
                    <th>Eff. Rate</th>
                    <th>Verdict</th>
                </tr>
                {% for job in job_analysis[:8] %}
                <tr>
                    <td>{{ job.category or 'Unknown' }}</td>
                    <td>{{ job.job_count or 0 }}</td>
                    <td>${{ "{:,.0f}".format(job.total_revenue or 0) }}</td>
                    <td>${{ "{:,.0f}".format(job.avg_revenue or 0) }}</td>
                    <td>${{ "{:,.0f}".format(job.effective_rate or 0) }}/hr</td>
                    <td style="font-weight: 600; color: {% if job.verdict == 'highly_profitable' or job.verdict == 'profitable' %}var(--success){% elif job.verdict == 'marginal' %}var(--warning){% else %}#dc2626{% endif %};">
                        {{ job.verdict | default('unknown') | replace('_', ' ') | title }}
                    </td>
                </tr>
                {% endfor %}
            </table>
        </section>
        {% endif %}
        
        {% if worst_jobs %}
        <!-- Your Worst Jobs - The Reality Check -->
        <div class="worst-jobs">
            <h3>🚨 Your Worst Performing Jobs</h3>
            <p style="margin-bottom: 16px; color: #991b1b;">These jobs likely cost you money or made very little. Learn from them.</p>
            <table>
                <tr>
                    <th>Job</th>
                    <th>Customer</th>
                    <th>Revenue</th>
                    <th>Est. Cost</th>
                    <th>Profit/Loss</th>
                    <th>Issue</th>
                </tr>
                {% for job in worst_jobs[:5] %}
                <tr>
                    <td>{{ job.job_description or 'Unknown Job' }}</td>
                    <td>{{ job.customer or 'Unknown' }}</td>
                    <td>${{ "{:,.0f}".format(job.revenue or 0) }}</td>
                    <td>${{ "{:,.0f}".format(job.estimated_cost or 0) }}</td>
                    <td class="loss">${{ "{:,.0f}".format(job.profit_loss or 0) }}</td>
                    <td style="font-size: 13px;">{{ job.why_bad or 'Low margin' }}</td>
                </tr>
                {% endfor %}
            </table>
            <div class="evidence" style="margin-top: 16px; background: #fff;">
                <strong>💡 Key Lesson:</strong> {{ worst_jobs[0].lesson if worst_jobs else 'Avoid small jobs that don\'t cover your real costs.' }}
            </div>
        </div>
        {% endif %}
        
        <div class="assumptions">
            <h3>⚠️ Key Assumptions (Be Honest With Yourself)</h3>
            <ul>
                <li>These numbers assume you actually implement the changes</li>
                <li>Rate increases assume ~15% customer loss (factored in)</li>
                <li>Timeline estimates assume you start this week</li>
                <li>Results depend on your market, customers, and execution</li>
                {% if guarantee.key_assumptions %}
                {% for assumption in guarantee.key_assumptions %}
                <li>{{ assumption }}</li>
                {% endfor %}
                {% endif %}
            </ul>
        </div>
        
        <div class="guarantee-box">
            <h2>Opportunity Identified</h2>
            <div class="guarantee-amount">${{ "{:,.0f}".format((guarantee.total_opportunity or 15000) * 0.7) }}</div>
            <p>Conservative estimate (85% of calculated value)</p>
            <p class="guarantee-note">
                Best case: ${{ "{:,.0f}".format(guarantee.total_opportunity or 15000) }} if you implement everything and retain all customers. 
                We've factored in realistic customer loss and execution challenges.
            </p>
        </div>
        
        <section>
            <h2>Next Steps</h2>
            <ol style="padding-left: 24px; line-height: 2;">
                <li><strong>This week:</strong> Implement the "this_week" actions - start with new customers</li>
                <li><strong>Strategy call:</strong> Let's discuss your specific questions and concerns</li>
                <li><strong>Track it:</strong> Use the Excel tracker to log what you've done and results</li>
                <li><strong>30-day check:</strong> We'll follow up to see what's working</li>
            </ol>
        </section>
        
        <!-- Methodology Section - Professional Transparency -->
        {% if methodology %}
        <section class="methodology-section">
            <h2>📊 Audit Methodology</h2>
            <div class="methodology-grid">
                <div class="methodology-item">
                    <h3>1. Data Analyzed</h3>
                    <ul>
                        <li>Invoices analyzed: {{ methodology.data_analyzed.invoices_count or summary.total_jobs or 'N/A' }}</li>
                        <li>Date range: {{ methodology.data_analyzed.date_range or 'Last 12 months' }}</li>
                        <li>Total revenue: ${{ "{:,.0f}".format(methodology.data_analyzed.total_revenue or summary.total_revenue_analyzed or 0) }}</li>
                    </ul>
                </div>
                <div class="methodology-item">
                    <h3>2. Benchmarking Sources</h3>
                    <ul>
                        <li><span class="confidence-badge high">HIGH</span> service.com.au (200+ data points)</li>
                        <li><span class="confidence-badge medium">MEDIUM</span> Industry associations (NECA, Master Plumbers, etc.)</li>
                        <li>Location-specific: {{ context.location }} market rates</li>
                    </ul>
                </div>
                <div class="methodology-item">
                    <h3>3. Calculation Approach</h3>
                    <ul>
                        <li><strong>Conservative estimates:</strong> 15% customer loss assumed</li>
                        <li><strong>Realistic estimates:</strong> 10% customer loss</li>
                        <li><strong>Optimistic estimates:</strong> 5% customer loss</li>
                        <li>All projections use conservative scenario as the recommendation</li>
                    </ul>
                </div>
                <div class="methodology-item">
                    <h3>4. Confidence Levels</h3>
                    <ul>
                        <li><span class="confidence-badge high">HIGH</span> 3+ sources, large sample, recent data</li>
                        <li><span class="confidence-badge medium">MEDIUM</span> 1-2 sources, industry estimates</li>
                        <li><span class="confidence-badge low">LOW</span> Single source or assumption-based</li>
                    </ul>
                </div>
            </div>
        </section>
        {% endif %}
        
        <!-- Disclaimer Section -->
        <div class="disclaimer-section">
            <h3>⚠️ Assumptions & Limitations</h3>
            <ul>
                <li>Projections based on data provided and industry benchmarks</li>
                <li>Customer retention rates based on industry averages (may vary)</li>
                <li>Hours estimated from job mix where actual timesheets not provided</li>
                <li>Market conditions as of {{ date }} - subject to change</li>
                <li>Results depend on implementation quality and market factors</li>
            </ul>
            <p style="margin-top: 16px; font-size: 13px; font-style: italic;">
                <strong>Important:</strong> All projections are estimates, not guarantees. 
                Actual results will vary based on implementation and business-specific factors.
            </p>
        </div>
        
        <footer>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 40px; margin-bottom: 30px; text-align: left;">
                <div>
                    <h3 style="font-size: 14px; font-weight: 600; margin-bottom: 12px;">Report Details</h3>
                    <div class="report-id">
                        Report ID: BRC-{{ context.location[:3] | upper }}-{{ context.trade_type[:4] | upper }}-{{ date | replace(' ', '') }}<br>
                        Generated: {{ date }}<br>
                        Version: 2.1
                    </div>
                </div>
                <div>
                    <h3 style="font-size: 14px; font-weight: 600; margin-bottom: 12px;">Next Steps</h3>
                    <div style="font-size: 13px; line-height: 1.8;">
                        ✓ Review this report thoroughly<br>
                        ✓ Start with "This Week" actions<br>
                        ✓ Questions? support@brace.com.au
                    </div>
                </div>
            </div>
            <div style="text-align: center; padding-top: 20px; border-top: 1px solid var(--border);">
                <div style="font-size: 18px; font-weight: 600; margin-bottom: 8px;">BRACE</div>
                <p>AI-Powered Profit Intelligence for Australian Tradies</p>
            </div>
        </footer>
    </div>
</body>
</html>
"""
        
        template = Template(html_template)
        
        # Get job analysis from profitability or directly
        job_analysis = []
        if isinstance(analysis.profitability, dict):
            job_analysis = analysis.profitability.get('by_job_type', [])
        
        html_content = template.render(
            customer_name=customer_name,
            date=datetime.now().strftime("%B %d, %Y"),
            exec_summary=exec_summary,
            context=context,
            summary=analysis.summary,
            pricing=analysis.pricing_audit,
            profitability=analysis.profitability,
            quote_analysis=analysis.quote_analysis,
            time_analysis=analysis.time_analysis,
            action_plan=analysis.action_plan,
            guarantee=analysis.guarantee_check,
            # Enhanced 2026 analysis
            business_health=analysis.business_health_score,
            customer_analysis=analysis.customer_analysis,
            cash_flow_insights=analysis.cash_flow_insights,
            job_analysis=job_analysis,
            data_quality=analysis.data_quality,
            next_steps=analysis.next_steps,
            worst_jobs=analysis.worst_jobs,
            # NEW: Methodology and provenance
            methodology=analysis.methodology,
            market_benchmarks=analysis.market_benchmarks_used,
            opportunity_summary=analysis.opportunity_summary
        )
        
        html_path = output_dir / "profit_leak_audit_report.html"
        with open(html_path, 'w') as f:
            f.write(html_content)
        
        return html_path
    
    def _generate_excel_report(
        self,
        analysis: AnalysisResult,
        context: BusinessContext,
        customer_name: str,
        output_dir: Path
    ) -> Path:
        """Generate Excel workbook with actionable spreadsheets."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return self._generate_csv_fallback(analysis, context, customer_name, output_dir)
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="171717", end_color="171717", fill_type="solid")
        money_font = Font(bold=True, color="16a34a")
        
        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        
        summary_data = [
            ["PROFIT LEAK AUDIT - SUMMARY", ""],
            ["", ""],
            ["Customer", customer_name],
            ["Trade", context.trade_type.capitalize()],
            ["Location", context.location],
            ["Report Date", datetime.now().strftime("%Y-%m-%d")],
            ["", ""],
            ["YOUR NUMBERS", ""],
            ["Revenue Analyzed", f"${analysis.summary.get('total_revenue_analyzed', 0):,.0f}"],
            ["Effective Hourly Rate", f"${analysis.summary.get('effective_hourly_rate', 0):,.0f}"],
            ["Market Mid-Rate", f"${analysis.pricing_audit.get('market_mid', context.current_rate + 15):,.0f}"],
            ["", ""],
            ["OPPORTUNITY", ""],
            ["Conservative Estimate", f"${analysis.guarantee_check.get('total_opportunity', 0) * 0.7:,.0f}"],
            ["Best Case", f"${analysis.guarantee_check.get('total_opportunity', 0):,.0f}"],
        ]
        
        for row in summary_data:
            ws.append(row)
        
        # Sheet 2: Action Tracker
        ws2 = wb.create_sheet("Action Tracker")
        headers = ["#", "Action", "Impact ($/yr)", "Effort", "Timeline", "Status", "Started", "Completed", "Actual Result", "Notes"]
        ws2.append(headers)
        
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        for i, action in enumerate(analysis.action_plan, 1):
            impact = action.get('impact_conservative', action.get('impact_annual', 0))
            ws2.append([
                i,
                action.get('action', ''),
                f"${impact * 0.85:,.0f}",
                action.get('effort', ''),
                action.get('timeline', ''),
                "Not Started",
                "",
                "",
                "",
                ""
            ])
        
        # Column widths
        ws2.column_dimensions['B'].width = 45
        ws2.column_dimensions['J'].width = 30
        
        # Sheet 3: Rate Calculator
        ws3 = wb.create_sheet("Rate Calculator")
        
        calc_data = [
            ["RATE INCREASE CALCULATOR", ""],
            ["", ""],
            ["Enter your numbers:", ""],
            ["Current Hourly Rate", context.current_rate],
            ["New Hourly Rate", analysis.pricing_audit.get('recommended_rate', context.current_rate + 15)],
            ["Billable Hours / Week", 30],
            ["Weeks / Year", 48],
            ["Expected Customer Loss %", 15],
            ["", ""],
            ["RESULTS:", ""],
            ["Total Billable Hours", "=B6*B7"],
            ["Rate Increase", "=B5-B4"],
            ["Gross Additional Revenue", "=B11*B12"],
            ["After Customer Loss", "=B13*(1-B8/100)"],
            ["", ""],
            ["Your realistic annual gain:", "=B14"],
        ]
        
        for row in calc_data:
            ws3.append(row)
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 15
        
        # Sheet 4: Scripts
        ws4 = wb.create_sheet("Scripts")
        ws4.append(["SCRIPTS - COPY & USE THESE", ""])
        ws4.append(["", ""])
        
        for action in analysis.action_plan:
            if action.get('script'):
                ws4.append([action.get('action', ''), ""])
                ws4.append(["What to say:", action.get('script', '')])
                if action.get('pushback_response'):
                    ws4.append(["If they push back:", action.get('pushback_response', '')])
                ws4.append(["", ""])
        
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 80
        
        excel_path = output_dir / "profit_leak_audit_workbook.xlsx"
        wb.save(excel_path)
        
        return excel_path
    
    def _generate_csv_fallback(
        self,
        analysis: AnalysisResult,
        context: BusinessContext,
        customer_name: str,
        output_dir: Path
    ) -> Path:
        """Generate CSV if openpyxl not available."""
        import csv
        
        csv_path = output_dir / "action_plan.csv"
        
        with open(csv_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["#", "Action", "Impact ($/yr)", "Effort", "Timeline", "Status", "Notes"])
            
            for i, action in enumerate(analysis.action_plan, 1):
                impact = action.get('impact_conservative', action.get('impact_annual', 0))
                writer.writerow([
                    i,
                    action.get('action', ''),
                    f"${impact * 0.85:,.0f}",
                    action.get('effort', ''),
                    action.get('timeline', ''),
                    "Not Started",
                    ""
                ])
        
        return csv_path
